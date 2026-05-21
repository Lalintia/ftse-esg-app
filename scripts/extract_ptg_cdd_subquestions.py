"""Extract sub-questions from PTG CDD Excel → indicator_subparts_ptg.json.

Source of truth: FTSE Russell CDD file (Indicator Question Code + Indicator Question).
Output schema matches indicator_subparts.json with source="cdd" + subsectors=["60101000"].
"""
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
CDD_PATH = Path(
    "/Users/alienmacbook/Desktop/OhmProject/PTG FTSE Report/563001436305883614_CDD-20250528-PTG Energy.xlsx"
)
INDICATORS_PATH = REPO_ROOT / "backend" / "data" / "ftse_indicators.json"
OUTPUT_PATH = REPO_ROOT / "backend" / "data" / "indicator_subparts_ptg.json"

PTG_SUBSECTOR = "60101000"
SHEET_NAME = "PTG Energy"


def load_indicator_metadata() -> dict[str, dict]:
    indicators = json.load(open(INDICATORS_PATH))
    return {ind["indicator_code"]: ind for ind in indicators}


def parse_subpart_letter(sub_code: str, sub_text: str) -> tuple[int | None, str | None]:
    """Extract sub-part number from EWT30_1 → (1, None) or detect leading 'a)' in text → (1, 'a')."""
    match = re.search(r"_(\d+)$", str(sub_code))
    num = int(match.group(1)) if match else None
    letter_match = re.match(r"^\s*([a-f])\)", str(sub_text)) if sub_text else None
    letter = letter_match.group(1) if letter_match else None
    return num, letter


def main() -> None:
    metadata = load_indicator_metadata()
    wb = openpyxl.load_workbook(CDD_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    output: dict[str, dict] = {}
    counts: dict[str, int] = defaultdict(int)
    skipped_missing_metadata: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        applicability = row[idx["Applicability Flag"]]
        if applicability != "YES":
            continue

        ind_code = row[idx["Indicator Code"]]
        sub_code = row[idx["Indicator Question Code"]]
        sub_text = row[idx["Indicator Question"]]

        if not ind_code or not sub_code:
            continue

        ind_code = str(ind_code).strip()
        sub_code = str(sub_code).strip()
        sub_text = str(sub_text).strip() if sub_text else ""

        meta = metadata.get(ind_code)
        if not meta:
            if ind_code not in skipped_missing_metadata:
                skipped_missing_metadata.append(ind_code)
            continue

        num, letter = parse_subpart_letter(sub_code, sub_text)

        entry = {
            "indicator_code": ind_code,
            "subpart_code": sub_code,
            "subpart_num": num,
            "subpart_letter": letter,
            "subpart_text": sub_text,
            "indicator_name": meta.get("indicator_name", ""),
            "theme_code": meta.get("theme_code", ""),
            "theme_name": meta.get("theme_name", ""),
            "data_type": meta.get("data_type", ""),
            "exposure_level": meta.get("exposure_level", ""),
            "is_core": meta.get("is_core", False),
            "type": meta.get("type", "core"),
            "marker": meta.get("marker", ""),
            "subsectors": [PTG_SUBSECTOR],
            "source": "cdd",
        }
        output[sub_code] = entry
        counts[ind_code] += 1

    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"✅ Extracted {len(output)} sub-questions from PTG CDD")
    print(f"   Indicators covered: {len(counts)}")
    print(f"   Output: {OUTPUT_PATH.relative_to(REPO_ROOT)}")
    if skipped_missing_metadata:
        print(
            f"⚠️  {len(skipped_missing_metadata)} indicators in CDD not found in "
            f"ftse_indicators.json: {skipped_missing_metadata[:10]}"
        )


if __name__ == "__main__":
    main()
