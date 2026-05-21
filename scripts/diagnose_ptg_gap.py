"""Diagnose PTG sub-indicator gap: indicator_subparts.json vs PTG CDD."""
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
SUBPARTS_PATH = REPO_ROOT / "backend" / "data" / "indicator_subparts.json"
CDD_PATH = Path("/Users/alienmacbook/Desktop/OhmProject/PTG FTSE Report/563001436305883614_CDD-20250528-PTG Energy.xlsx")

sys.path.insert(0, str(REPO_ROOT / "backend"))
from app.utils.sector_themes import get_applicable_themes

PTG_SUBSECTOR = "60101000"


def load_cdd_counts() -> tuple[dict[str, int], dict[str, list[str]]]:
    """Return (indicator_code -> sub_question_count, indicator_code -> [sub_codes]) for PTG YES rows."""
    wb = openpyxl.load_workbook(CDD_PATH, data_only=True)
    ws = wb["PTG Energy"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    counts: dict[str, int] = defaultdict(int)
    codes: dict[str, list[str]] = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        applicability = row[idx["Applicability Flag"]]
        if applicability != "YES":
            continue
        ind_code = row[idx["Indicator Code"]]
        sub_code = row[idx["Indicator Question Code"]]
        if not ind_code or not sub_code:
            continue
        counts[ind_code] += 1
        codes[ind_code].append(sub_code)

    return dict(counts), dict(codes)


def load_subparts_counts() -> dict[str, int]:
    """Count sub-indicators per indicator_code from indicator_subparts.json, filtered for PTG."""
    data = json.load(open(SUBPARTS_PATH))
    themes = get_applicable_themes(PTG_SUBSECTOR)
    applicable_theme_names = {
        t["theme"] for t in themes if t.get("indicators_applicable", True)
    }

    counts: dict[str, int] = defaultdict(int)
    for sub_code, entry in data.items():
        if entry["theme_name"] not in applicable_theme_names:
            continue

        subsectors = entry.get("subsectors", [])
        ind_type = entry.get("type", "core")
        if ind_type in ("specific", "geography"):
            if PTG_SUBSECTOR not in subsectors:
                continue

        counts[entry["indicator_code"]] += 1

    return dict(counts)


def main() -> None:
    cdd_counts, cdd_codes = load_cdd_counts()
    our_counts = load_subparts_counts()

    all_indicators = sorted(set(cdd_counts) | set(our_counts))

    cdd_total = sum(cdd_counts.values())
    our_total = sum(our_counts.values())
    print(f"=== PTG Sub-indicator Gap Report ===")
    print(f"CDD total sub-questions: {cdd_total}")
    print(f"Our total sub-indicators: {our_total}")
    print(f"Gap: {our_total - cdd_total}")
    print(f"CDD indicators (YES): {len(cdd_counts)}")
    print(f"Our indicators: {len(our_counts)}")
    print()

    print(f"{'Indicator':<10} {'Ours':>5} {'CDD':>5} {'Diff':>6}  Status")
    print("-" * 50)
    gap_perf = 0
    gap_qual = 0
    gap_other = 0
    extra_in_ours = []
    missing_in_ours = []
    matched = 0

    for ind in all_indicators:
        ours = our_counts.get(ind, 0)
        cdd = cdd_counts.get(ind, 0)
        diff = ours - cdd
        if diff == 0 and ours > 0:
            matched += 1
            continue
        if cdd == 0:
            extra_in_ours.append(ind)
            status = "EXTRA (not in PTG CDD)"
        elif ours == 0:
            missing_in_ours.append(ind)
            status = "MISSING in ours"
        else:
            status = "COUNT MISMATCH"
        marker = ""
        if cdd >= 7 and ours == 1:
            marker = " [PERFORMANCE?]"
            gap_perf += (cdd - ours)
        elif cdd > ours:
            gap_qual += (cdd - ours)
        else:
            gap_other += (ours - cdd)
        print(f"{ind:<10} {ours:>5} {cdd:>5} {diff:>+6}  {status}{marker}")

    print()
    print(f"Matched (count equal, both >0): {matched}")
    print(f"Extra in ours (not in PTG CDD applicable): {len(extra_in_ours)}")
    print(f"Missing in ours: {len(missing_in_ours)}")
    print(f"Gap categorization:")
    print(f"  - Performance-style (cdd>=7, ours=1): -{gap_perf}")
    print(f"  - Qualitative miss (cdd>ours): -{gap_qual}")
    print(f"  - Other / extra: +{gap_other}")


if __name__ == "__main__":
    main()
