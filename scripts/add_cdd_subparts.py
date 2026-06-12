"""Add sub-questions found in the PTG CDD but missing from indicator_subparts.json.

The CDD (Company Data Download) is FTSE's official per-company workbook; its
"Indicator Question Code" column reflects the real data-model structure
(e.g. EPR18 has 7 sub-questions, our KB only had 6). This script copies the
missing question codes + English question text into indicator_subparts.json,
inheriting all indicator-level metadata from an existing sibling subpart.

Idempotent: codes already present are skipped.

Usage:
    python3 scripts/add_cdd_subparts.py
"""

import json
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
SUBPARTS_PATH = REPO_ROOT / "backend" / "data" / "indicator_subparts.json"
CDD_PATH = Path(
    "/Users/alienmacbook/Desktop/OhmProject/PTG FTSE Report/"
    "563001436305883614_CDD-20250528-PTG Energy.xlsx"
)

sys.path.insert(0, str(REPO_ROOT / "scripts"))
from verify_against_refs import ref_ptg_from_cdd, system_applicable_subparts  # noqa: E402


def load_cdd_questions() -> dict[str, str]:
    """Return {question_code: question_text} for every row in the CDD."""
    wb = openpyxl.load_workbook(CDD_PATH, data_only=True)
    ws = wb["PTG Energy"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    questions: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[idx["Indicator Question Code"]] or "").strip()
        text = str(row[idx["Indicator Question"]] or "").strip()
        if code and text:
            questions[code] = text
    return questions


def main() -> None:
    subparts = json.load(open(SUBPARTS_PATH, encoding="utf-8"))
    questions = load_cdd_questions()

    ref = ref_ptg_from_cdd()
    ours = set(system_applicable_subparts("60101000"))
    missing = sorted(set(ref["sub_codes"]) - ours - set(subparts))

    added, skipped = 0, []
    for sub_code in missing:
        ind_code, _, num = sub_code.rpartition("_")

        # Inherit indicator metadata from any existing sibling subpart
        sibling = next(
            (v for k, v in subparts.items() if k.startswith(f"{ind_code}_")),
            None,
        )
        if sibling is None:
            skipped.append(sub_code)
            continue

        entry = dict(sibling)
        entry["subpart_code"] = sub_code
        entry["subpart_num"] = int(num)
        entry["subpart_letter"] = ""
        entry["subpart_text"] = questions[sub_code]
        subparts[sub_code] = entry
        added += 1

    with open(SUBPARTS_PATH, "w", encoding="utf-8") as f:
        json.dump(subparts, f, ensure_ascii=False, indent=2)

    print(f"added {added}, skipped (no sibling) {skipped or 'none'}")
    print(f"total subparts now: {len(subparts)}")


if __name__ == "__main__":
    main()
