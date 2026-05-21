"""Verify our system against ground-truth references — Tidlor (Ake PDF) + PTG (CDD).

Rules:
1. Our system derives applicable indicators + sub-indicators using ONLY:
   - indicator_subsector_mapping.json
   - sector_themes.py
   - indicator_subparts.json (description-based parsing)
2. REF documents are READ-ONLY ground truth for comparison.
3. NO data from REF flows back into the system (no override files used).
"""
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pdfplumber

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "backend"))

from app.utils.sector_themes import get_applicable_themes

DATA_DIR = REPO_ROOT / "backend" / "data"
SUBPARTS = json.loads((DATA_DIR / "indicator_subparts.json").read_text())
MAPPING = json.loads((DATA_DIR / "indicator_subsector_mapping.json").read_text())
INDICATORS = {
    ind["indicator_code"]: ind
    for ind in json.loads((DATA_DIR / "ftse_indicators.json").read_text())
}

TIDLOR_PDF = Path(
    "/Users/alienmacbook/Desktop/OhmProject/ESG Knowledge/"
    "Tidlor_FTSE_Gap_Analysis (By Ake).pdf"
)
PTG_CDD = Path(
    "/Users/alienmacbook/Desktop/OhmProject/PTG FTSE Report/"
    "563001436305883614_CDD-20250528-PTG Energy.xlsx"
)


# ---------------------------------------------------------------------------
# OUR SYSTEM — derive applicable indicators + sub-indicators per subsector
# ---------------------------------------------------------------------------

def system_applicable_indicators(subsector_code: str) -> set[str]:
    """Indicators our system thinks apply to a subsector.

    Filter: theme applicable AND (core/performance OR subsector in subsectors).
    """
    themes = get_applicable_themes(subsector_code)
    applicable_themes = {
        t["theme"] for t in themes if t.get("indicators_applicable", True)
    }

    result = set()
    for code, info in MAPPING.items():
        meta = INDICATORS.get(code)
        if not meta or meta.get("theme_name") not in applicable_themes:
            continue
        ind_type = info.get("type", "core")
        if ind_type in ("core", "performance"):
            if subsector_code in info.get("exclude_subsectors", []):
                continue
            result.add(code)
        elif ind_type in ("specific", "geography"):
            if subsector_code in info.get("subsectors", []):
                result.add(code)
    return result


def system_applicable_subparts(subsector_code: str) -> set[str]:
    """Sub-indicators our system derives via description-based parsing + filter."""
    themes = get_applicable_themes(subsector_code)
    applicable_themes = {
        t["theme"] for t in themes if t.get("indicators_applicable", True)
    }
    applicable_indicators = system_applicable_indicators(subsector_code)

    result = set()
    for sub_code, entry in SUBPARTS.items():
        if entry["indicator_code"] not in applicable_indicators:
            continue
        if entry["theme_name"] not in applicable_themes:
            continue
        result.add(sub_code)
    return result


# ---------------------------------------------------------------------------
# REF EXTRACTION — read-only, NOT fed back into system
# ---------------------------------------------------------------------------

def ref_tidlor_from_ake_pdf() -> dict:
    """Extract reference info from Ake PDF — counts + 62 gap codes."""
    with pdfplumber.open(TIDLOR_PDF) as pdf:
        full_text = "\n".join((p.extract_text() or "") for p in pdf.pages)

    # Headline numbers from page 3
    counts = {"indicators": 120, "sub_indicators": 219}

    # Pass 1: primary tokens "ECC03_1", "ECC74_1-2", "ECC75_1, 75_3"
    raw_tokens = re.findall(
        r"\b([EGS][A-Z]{2}\d{1,3})_(\d+(?:[-,]\s*(?:\d+_)?\d+)*)",
        full_text,
    )
    gap_codes: set[str] = set()
    last_prefix: str = ""
    for prefix, suffix in raw_tokens:
        last_prefix = prefix
        for piece in re.split(r"[,\s]+", suffix):
            piece = piece.strip()
            if not piece:
                continue
            if "-" in piece:
                a, b = piece.split("-", 1)
                try:
                    for n in range(int(a), int(b) + 1):
                        gap_codes.add(f"{prefix}_{n}")
                except ValueError:
                    continue
            else:
                if "_" in piece:
                    continue
                gap_codes.add(f"{prefix}_{piece}")

    # Pass 2: continuation tokens "75_3" / "14_3" — e.g. "ECC75_1, 75_3" means
    # ECC75_1 and ECC75_3. The "75_3" part uses only the indicator number, not the
    # theme prefix. For each line, find ALL primary codes, then scan for continuations
    # using the same indicator number.
    for line in full_text.split("\n"):
        for theme_prefix, ind_num in re.findall(r"\b([EGS][A-Z]{2})(\d{1,3})_\d+", line):
            ind_code = theme_prefix + ind_num
            # Look for "NN_M" patterns where NN == ind_num (not preceded by letters)
            for cont in re.findall(rf"(?<![A-Za-z]){re.escape(ind_num)}_(\d+)", line):
                gap_codes.add(f"{ind_code}_{cont}")

    return {"counts": counts, "gap_codes": sorted(gap_codes)}


def ref_ptg_from_cdd() -> dict:
    """Extract reference indicators + sub-question codes from PTG CDD."""
    wb = openpyxl.load_workbook(PTG_CDD, data_only=True)
    ws = wb["PTG Energy"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    indicators: set[str] = set()
    sub_codes: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Applicability Flag"]] != "YES":
            continue
        ind = row[idx["Indicator Code"]]
        sub = row[idx["Indicator Question Code"]]
        if ind:
            indicators.add(str(ind).strip())
        if sub:
            sub_codes.add(str(sub).strip())
    return {"indicators": sorted(indicators), "sub_codes": sorted(sub_codes)}


# ---------------------------------------------------------------------------
# COMPARISON
# ---------------------------------------------------------------------------

def diff_block(label: str, ours: set, theirs: set) -> str:
    matched = ours & theirs
    only_ours = ours - theirs
    only_theirs = theirs - ours
    lines = [
        f"  {label}",
        f"    Ours:     {len(ours)}",
        f"    Ref:      {len(theirs)}",
        f"    Matched:  {len(matched)}",
        f"    Only ours: {len(only_ours)}" + (
            f"  → {sorted(only_ours)[:8]}{'…' if len(only_ours) > 8 else ''}"
            if only_ours else ""
        ),
        f"    Only ref:  {len(only_theirs)}" + (
            f"  → {sorted(only_theirs)[:8]}{'…' if len(only_theirs) > 8 else ''}"
            if only_theirs else ""
        ),
    ]
    return "\n".join(lines)


def verify_tidlor() -> None:
    print("=" * 70)
    print("TIDLOR  —  ICB 30201020 (Consumer Lending)")
    print("=" * 70)

    sub_code = "30201020"
    sys_indicators = system_applicable_indicators(sub_code)
    sys_subparts = system_applicable_subparts(sub_code)
    ref = ref_tidlor_from_ake_pdf()

    print("\n[Industry Classification]")
    print("  System:  Financials → Consumer Lending (30201020)")
    print("  REF:     Financials → Specialized Consumer Services (Ake PDF)")

    print("\n[Themes applicable]")
    themes = get_applicable_themes(sub_code)
    applicable = [t["theme"] for t in themes if t.get("indicators_applicable", True)]
    print(f"  System: {sorted(applicable)}")
    print(
        "  REF:    ['Anti-Corruption', 'Climate Change', 'Corporate Governance', "
        "'Customer Responsibility', 'Human Rights & Community', "
        "'Labour Standards', 'Risk Management']"
    )

    print("\n[Indicator counts]")
    print(f"  System:  {len(sys_indicators)}")
    print(f"  REF:     {ref['counts']['indicators']}")
    delta = len(sys_indicators) - ref["counts"]["indicators"]
    print(f"  Delta:   {delta:+d}  {'✅' if delta == 0 else '⚠️'}")

    print("\n[Sub-indicator counts]")
    print(f"  System:  {len(sys_subparts)}")
    print(f"  REF:     {ref['counts']['sub_indicators']}")
    delta = len(sys_subparts) - ref["counts"]["sub_indicators"]
    print(f"  Delta:   {delta:+d}  ({delta/ref['counts']['sub_indicators']*100:+.1f}%)")

    print("\n[Sub-indicator delta — Phase 31 expansion impact]")
    from collections import Counter
    ind_counts: Counter = Counter(SUBPARTS[k]["indicator_code"] for k in sys_subparts if k in SUBPARTS)
    phase31_expanded = ["GAC03", "GAC04", "GAC05", "GAC07", "GAC08", "GRM04", "GRM05", "SLS03", "SLS16"]
    print(f"  Ake REF counts each expanded indicator as 1 row; Phase 31 split these:")
    for ind in phase31_expanded:
        cnt = ind_counts.get(ind, 0)
        if cnt > 0:
            print(f"    {ind}: {cnt} sub-parts (REF counts as 1)")
    print(f"  Known limitation — Ake PDF uses single-row format, not expandable sub-parts.")

    print("\n[Gap codes from Ake PDF — does our system mark these as applicable?]")
    ref_gaps = set(ref["gap_codes"])
    # B-BBEE (South Africa) noted as "ไม่เกี่ยวข้อง" in PDF — exclude
    south_africa_codes = {c for c in ref_gaps if c.startswith("SLS27_")}
    ref_gaps_clean = ref_gaps - south_africa_codes
    coverage = ref_gaps_clean & sys_subparts
    missing = ref_gaps_clean - sys_subparts
    print(f"  Extracted from PDF: {len(ref_gaps)} (incl. {len(south_africa_codes)} SLS27 South Africa excluded)")
    print(f"  Reviewable gaps:    {len(ref_gaps_clean)}")
    print(f"  Found in our list:  {len(coverage)} ({len(coverage)/len(ref_gaps_clean)*100:.0f}%)")
    if missing:
        print(f"  Missing from our list: {sorted(missing)}")


def verify_ptg() -> None:
    print("\n" + "=" * 70)
    print("PTG  —  ICB 60101000 (Integrated Oil & Gas)")
    print("=" * 70)

    sub_code = "60101000"
    sys_indicators = system_applicable_indicators(sub_code)
    sys_subparts = system_applicable_subparts(sub_code)
    ref = ref_ptg_from_cdd()
    ref_indicators = set(ref["indicators"])
    ref_subs = set(ref["sub_codes"])

    print("\n[Industry Classification]")
    print("  System:  Energy → Oil, Gas & Coal → Integrated Oil & Gas (60101000)")
    print("  REF:     Oil & Gas → Oil & Gas Producers → Integrated Oil & Gas (FTSE CPC)")

    print("\n[Themes applicable]")
    themes = get_applicable_themes(sub_code)
    applicable = [t["theme"] for t in themes if t.get("indicators_applicable", True)]
    print(f"  System: {sorted(applicable)}")
    # From CDD: themes present with at least one YES indicator
    ref_themes = {INDICATORS[i].get("theme_name") for i in ref_indicators if i in INDICATORS}
    print(f"  REF:    {sorted(ref_themes)}")

    print()
    print(diff_block("Indicators", sys_indicators, ref_indicators))
    print()
    print(diff_block("Sub-indicators", sys_subparts, ref_subs))


def main() -> None:
    verify_tidlor()
    verify_ptg()


if __name__ == "__main__":
    main()
